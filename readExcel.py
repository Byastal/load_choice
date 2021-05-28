import xlwings as xw
import json
import time
import requests

def load_choice(file_path,url):
    # app=xw.App(visible=True,add_book=False)
    # excel_client = app.books.open(file_path)
    app = xw.App(visible=True,add_book=False)

    print("app created!")

    excel_client=app.books.open(r"D:\load_choice\model.xlsm")
    print(1)
    sheet = excel_client.sheets("吃贴水策略")
    print(sheet)

    date = time.strftime("%Y-%m-%d ", time.localtime()) 
    json_data = {}
    
    monitor_pos = [{'stock':'F','stock_range':(1,3),'contract':'BCDE','contract_range':(1,10)},
            {'stock':'F','stock_range':(13,15),'contract':'BCDE','contract_range':(13,22)},
            {'stock':'M','stock_range':(1,3),'contract':'IJKL','contract_range':(1,10)},
    ]
    for pos in monitor_pos:
        # tmp = {}
        stock_name = sheet.range(pos['stock']+str(pos['stock_range'][0])).value
        json_data[stock_name]={'update_time':date+str(sheet.range(pos['stock']+str(pos['stock_range'][0]+1)).value),
                        'latest_price':sheet.range(pos['stock']+str(pos['stock_range'][0]+2)).value  
                        }
        for i in pos['contract']:
            start= pos['contract_range'][0]
            topic_name = sheet.range(i+str(start)).value
            update_time = sheet.range(i+str(start+1)).value
            latest_price = sheet.range(i+str(start+2)).value
            contract_year = sheet.range(i+str(start+3)).value
            current_contract_month = sheet.range(i+str(start+4)).value
            delivery_day = sheet.range(i+str(start+5)).value
            premium = sheet.range(i+str(start+6)).value
            premium_ratio = sheet.range(i+str(start+7)).value
            days_before_expiration = sheet.range(i+str(start+8)).value
            annualized_premium_ratio = sheet.range(i+str(start+9)).value
            json_data[topic_name]={'update_time':date+str(update_time),
                            'latest_price':latest_price,
                            # 'contract_year':contract_year,
                            # 'current_contract_month':current_contract_month,
                            # 'delivery_day':delivery_day,
                            # 'premium':premium,
                            # 'premium_ratio':premium_ratio,
                            # 'days_before_expiration':days_before_expiration,
                            # 'annualized_premium_ratio':annualized_premium_ratio
                            }
        # json_data.append(tmp)
    print(json_data)
    json_data = json.dumps(json_data,ensure_ascii=False).encode("utf-8")
    # return json_data
    response = requests.post(url,data = json_data)
    return response.status_code
    # return 200

def test_load(file_path):
    from openpyxl import load_workbook  # 读到了VBA公式代码

    # wb = load_workbook(file_path,keep_vba=True)
    # ws = wb.worksheets[1]
    # print(ws.cell(row=2, column=3).value )
    
    import xlrd  # 导入库,直接失败
    # 打开文件
    # xlsx = xlrd.open_workbook(file_path)
    # # 查看所有sheet列表
    # sheet1 = xlsx.sheets()[0]
    # print(sheet1.row(2)[3].value )



if __name__ == "__main__":
    file_path = "D://load_choice//model.xlsm"
    url ="http://124.71.113.79:80/api/financial/real_time_data"
    print(file_path)
    # test_load(file_path)
    data = load_choice(file_path,url)
    print(data)
    # app = xw.App(visible=True, add_book=False)
    # excel_client = app.books.open(file_path)
    
    # time.sleep(10)
    # while True:
    # data = load_choice(file_path,url)
    # print(data)
    # response = requests.post(url,data = data.encode("utf-8"))
    # print(response.status_code)